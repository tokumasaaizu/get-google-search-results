import json                                 # 返却された検索結果の読み取りにつかう
from googleapiclient.discovery import build # APIへのアクセスにつかう
import chromedriver_binary
from selenium import webdriver
import lxml.html as lx
import re
import requests
from html.parser import HTMLParser
from urllib import request
from bs4 import BeautifulSoup
import openpyxl as op
import time
import logging
import subprocess
import traceback
import logging.handlers
# ログローテーション/フォーマットの設定
format_base = '%(asctime)s [%(levelname)s] [Thread-%(thread)d] %(filename)s:(%(lineno)d).%(funcName)s -%(message)s'
format = logging.Formatter(format_base)
output_log = './logs/'
filename = output_log + 'screip-case.log'
fh=logging.handlers.TimedRotatingFileHandler(filename=filename, when='MIDNIGHT', backupCount=8, encoding='utf-8')
fh.setLevel(logging.INFO)
fh.setFormatter(format)
ch = logging.StreamHandler()
ch.setLevel(logging.ERROR)
ch.setFormatter(format)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(fh)
logger.addHandler(ch)

result_append = []
# カスタム検索エンジンID
CUSTOM_SEARCH_ENGINE_ID = "xxxxxxxx"
# API キー
API_KEY = "xxxxxxxxxx"

# APIにアクセスして結果をもらってくるメソッド
def get_search_results(query):
   # APIでやりとりするためのリソースを構築
   # 詳細: https://googleapis.github.io/google-api-python-client/docs/epy/googleapiclient.discovery-pysrc.html#build
   search = build(
       "customsearch", 
       "v1", 
       developerKey = API_KEY
   )
   
   # Google Custom Search から結果を取得
   # 詳細: https://developers.google.com/custom-search/v1/reference/rest/v1/cse/list
   result = search.cse().list(
       q = query,
       cx = CUSTOM_SEARCH_ENGINE_ID,
       lr = 'lang_ja',
       num = 10,
       start = 1
   ).execute()
   # 受け取ったjsonをそのまま返却
   return result


# 検索結果の情報をSearchResultに格納してリストで返す
def summarize_search_results(result):
   # 結果のjsonから検索結果の部分を取り出しておく
   result_items_part = result['items']
   # 抽出した検索結果の情報はこのリストにまとめる
   result_items = []
   # 今回は (start =) 1 個目の結果から (num =) 10 個の結果を取得した
   for i in range(0, 10):
       # i番目の検索結果の部分
       result_item = result_items_part[i]
       # i番目の検索結果からそれぞれの属性の情報をResultクラスに格納して
       # result_items リストに追加する
       result_items.append(result_item['title'])
       result_items.append(result_item['link'])
       result_items.append(result_item['snippet'])
   # 結果を格納したリストを返却
   return result_items
       
# 検索結果の情報を格納するクラス
class SearchResult:
   def __init__(self, title, url, snippet, rank):
       self.title = title
       self.url = url
       self.snippet = snippet
       self.rank = rank
   def __str__(self):
       # コマンドライン上での表示形式はご自由にどうぞ
       return "[title] " + self.title + "\n\t[url] " + self.url + "\n\t[snippet] " + self.snippet + "\n\t[rank] " + str(self.rank)

# メインプロセス       
if __name__ == '__main__':
    loop_count = 1
    #ブックの読み込み
    wb_in = op.load_workbook('banar.xlsx')
    ws_in = wb_in.worksheets[0]
    #ワークブック(Excelファイル)の新規作成
    wb_out = op.Workbook()
    #シートを取得し、名前を変える
    sheet = wb_out.active
    sheet.title = "output"
    cells = sheet.cell(row=10000, column=50)
    ws_out = wb_out['output']  # ワークシートを指定
    ws_out = wb_out.active
    # アウトプットシートのタイトル
    ws_out.cell(row=1, column=1).value = '番号'
    ws_out.cell(row=1, column=2).value = '番号'
    ws_out.cell(row=1, column=3).value = '都道府県'
    ws_out.cell(row=1, column=4).value = '市'
    ws_out.cell(row=1, column=5).value = '区'
    ws_out.cell(row=1, column=6).value = '募集url'
    ws_out.cell(row=1, column=7).value = 'トップページurl'
    ws_out.cell(row=1, column=8).value = 'タイトル'
    ws_out.cell(row=1, column=9).value = '概要文'
    while loop_count < 515:
        num = ws_in.cell(loop_count+1,1).value
        num_2 = ws_in.cell(loop_count+1,2).value
        todoufuken = ws_in.cell(loop_count+1,3).value
        city = ws_in.cell(loop_count+1,4).value
        ku = ''
        if ws_in.cell(loop_count+1,5).value != '':
            ku = ws_in.cell(loop_count+1,5).value
        try:
            keyword = ws_in.cell(loop_count+1,7).value
            query = keyword
            # APIから検索結果を取得
            # result には 返却されたjsonが入る
            result = get_search_results(query)
            # 検索結果情報からタイトル, URL, スニペット, 検索結果の順位を抽出してまとめる
            result_items_list = summarize_search_results(result) # result_items_list には SearchResult のリストが入る
            # 検索結果の情報をシートに出力
            #print(result_items_list)
            count = (loop_count - 1) * 10
            count = count + 2
            for i in range(0, 10):
                # シートに記入
                search_count = i*3
                ws_out.cell(row=count+i, column=1).value = num
                ws_out.cell(row=count+i, column=2).value = num_2
                ws_out.cell(row=count+i, column=3).value = todoufuken
                ws_out.cell(row=count+i, column=4).value = city
                ws_out.cell(row=count+i, column=5).value = ku
                #タイトル
                ws_out.cell(row=count+i, column=8).value = result_items_list[0+search_count]
                #募集URL
                ws_out.cell(row=count+i, column=6).value = result_items_list[1+search_count]
                #トップURL
                top_left = result_items_list[1+search_count].split('/')[0]
                top_right = result_items_list[1+search_count].split('/')[2]
                ws_out.cell(row=count+i, column=7).value = top_left + '//' + top_right
                #概要文
                ws_out.cell(row=count+i, column=9).value = result_items_list[2+search_count]
                #print(result_items_list[2+i])
            loop_count += 1
            time.sleep(2)
        except:
            for err in traceback.format_exc().split('\n'):
                logger.error(err)
        finally:
            pass
    wb_out.save('output-sample.xls')
