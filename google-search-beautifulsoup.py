# -*- coding: utf-8 -*-

import chromedriver_binary
from selenium import webdriver
import lxml.html as lx
import re
import requests
from html.parser import HTMLParser
from urllib import request
from bs4 import BeautifulSoup
import openpyxl as op
import time
import requests_cache
requests_cache.install_cache(cache_name='test_cache', backend='sqlite', expire_after=60*120)
import logging
import subprocess
import traceback
import logging.handlers
# ログローテーション/フォーマットの設定
format_base = '%(asctime)s [%(levelname)s] [Thread-%(thread)d] %(filename)s:(%(lineno)d).%(funcName)s -%(message)s'
format = logging.Formatter(format_base)
output_log = './logs/'
filename = output_log + 'screip-case.log'
fh=logging.handlers.TimedRotatingFileHandler(filename=filename, when='MIDNIGHT', backupCount=8, encoding='utf-8')
fh.setLevel(logging.INFO)
fh.setFormatter(format)
ch = logging.StreamHandler()
ch.setLevel(logging.ERROR)
ch.setFormatter(format)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(fh)
logger.addHandler(ch)


loop_count = 1
#ブックの読み込み
wb_in = op.load_workbook('banar.xlsx')
ws_in = wb_in.worksheets[0]
#ワークブック(Excelファイル)の新規作成
wb_out = op.Workbook()
#シートを取得し、名前を変える
sheet = wb_out.active
sheet.title = "output"
cells = sheet.cell(row=10000, column=50)
ws_out = wb_out['output']  # ワークシートを指定
ws_out = wb_out.active
# アウトプットシートのタイトル
ws_out.cell(row=1, column=1).value = '番号'
ws_out.cell(row=1, column=2).value = '番号'
ws_out.cell(row=1, column=3).value = '都道府県'
ws_out.cell(row=1, column=4).value = '市'
ws_out.cell(row=1, column=5).value = '区'
ws_out.cell(row=1, column=6).value = '募集url'
ws_out.cell(row=1, column=7).value = 'トップページurl'
ws_out.cell(row=1, column=8).value = 'タイトル'
ws_out.cell(row=1, column=9).value = '概要文'
while loop_count < 2:
    num = ws_in.cell(loop_count+1,1).value
    num_2 = ws_in.cell(loop_count+1,2).value
    todoufuken = ws_in.cell(loop_count+1,3).value
    city = ws_in.cell(loop_count+1,4).value
    ku = ''
    if ws_in.cell(loop_count+1,5).value != '':
        ku = ws_in.cell(loop_count+1,5).value
    # キーワード抽出
    try:
        keyword = ws_in.cell(loop_count+1,7).value
        search_query = keyword
        r = requests.get('https://www.google.co.jp/search?hl=jp&gl=JP&num=10&q=' + search_query)
        print(r)
        soup = BeautifulSoup(r.text, "html.parser")
        #print(soup)
        search_site_list = soup.select('div.kCrYT > a')
        print(search_site_list)
        for data in search_site_list:
            url = data.attrs['href']
            url_del = url.replace('/url?q=', '')
            url_end = url_del.split('&')
        for data_url in search_site_list:
            title = data_url.select('h3.zBAuLc')
    except:
        for err in traceback.format_exc().split('\n'):
            logger.error(err)
    finally:
        pass

    # シートに記入
    ws_out.cell(row=loop_count+1, column=1).value = num
    ws_out.cell(row=loop_count+1, column=2).value = num_2
    ws_out.cell(row=loop_count+1, column=3).value = todoufuken
    ws_out.cell(row=loop_count+1, column=4).value = city
    ws_out.cell(row=loop_count+1, column=5).value = ku
    ws_out.cell(row=loop_count+1, column=6).value = url_end[0]
    ws_out.cell(row=loop_count+1, column=8).value = title[0].contents[0].contents[0]
    loop_count += 1
    time.sleep(3)
wb_out.save('output-sample.xls')
